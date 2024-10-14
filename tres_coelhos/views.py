from django.shortcuts import render, redirect
from django.utils import timezone
import random
from django.http import HttpResponse
# from openpyxl import Workbook  # Para gerar o Excel
from openpyxl import load_workbook
import qrcode  # Para gerar o QR Code
from io import BytesIO  # Para manipular imagens em memória
from .models import Apartamento, Vaga, Sorteio, SorteioDupla, DuplaApartamentos


def tres_coelhos_sorteio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()
        print("Sorteios anteriores apagados.")

        # Obter apartamentos PNE, Idosos e Normais, excluindo os com apenas_dupla=True onde não aplicável
        apartamentos_pne_apenas_livre = list(Apartamento.objects.filter(is_pne=True, apenas_livre=True, apenas_dupla=False))
        apartamentos_pne_sem_restricao = list(Apartamento.objects.filter(is_pne=True, apenas_livre=False, apenas_dupla=False))
        apartamentos_idoso = list(Apartamento.objects.filter(is_idoso=True, apenas_dupla=False))
        apartamentos_normais = list(Apartamento.objects.filter(is_pne=False, is_idoso=False, apenas_dupla=False))

        print(f"Apartamentos PNE (apenas livres): {len(apartamentos_pne_apenas_livre)}")
        print(f"Apartamentos PNE (sem restrições): {len(apartamentos_pne_sem_restricao)}")
        print(f"Apartamentos Idosos: {len(apartamentos_idoso)}")
        print(f"Apartamentos Normais: {len(apartamentos_normais)}")

        # Obter vagas do banco de dados separadas por especialidade e tipo
        vagas_pne_livres = list(Vaga.objects.filter(especial='PNE', tipo='LIVRE'))
        vagas_pne_duplas = list(Vaga.objects.filter(especial='PNE', tipo='DUPLA'))
        vagas_idosos = list(Vaga.objects.filter(especial='IDOSO', tipo='LIVRE'))
        vagas_idosos_duplas = list(Vaga.objects.filter(especial='IDOSO', tipo='DUPLA'))
        vagas_livres = list(Vaga.objects.filter(especial='NORMAL', tipo='LIVRE'))

        print(f"Vagas PNE Livres: {len(vagas_pne_livres)}, Vagas PNE Duplas: {len(vagas_pne_duplas)}")
        print(f"Vagas Idosos Livres: {len(vagas_idosos)}, Vagas Idosos Duplas: {len(vagas_idosos_duplas)}")
        print(f"Vagas Livres: {len(vagas_livres)}")

        # Apartamentos que já ganharam vagas especiais (para não concorrerem nas vagas livres)
        apartamentos_com_vaga = []

        # **Sorteio prioritário para PNE com apenas_livre=True (vagas livres PNE)**
        random.shuffle(vagas_pne_livres)  # Aleatoriza as vagas PNE livres
        for vaga in vagas_pne_livres:
            if apartamentos_pne_apenas_livre:
                apartamento_escolhido = random.choice(apartamentos_pne_apenas_livre)
                Sorteio.objects.create(apartamento=apartamento_escolhido, vaga=vaga)
                print(f"Sorteado (PNE - apenas_livre): {apartamento_escolhido.numero} para a vaga PNE Livre {vaga.numero}")
                apartamentos_pne_apenas_livre.remove(apartamento_escolhido)
                apartamentos_com_vaga.append(apartamento_escolhido)
            else:
                # Se todos apartamentos com restrição já foram sorteados, preencher as vagas restantes com apartamentos sem restrição
                if apartamentos_pne_sem_restricao:
                    apartamento_escolhido = random.choice(apartamentos_pne_sem_restricao)
                    Sorteio.objects.create(apartamento=apartamento_escolhido, vaga=vaga)
                    print(f"Sorteado (PNE - sem restrição): {apartamento_escolhido.numero} para a vaga PNE Livre {vaga.numero}")
                    apartamentos_pne_sem_restricao.remove(apartamento_escolhido)
                    apartamentos_com_vaga.append(apartamento_escolhido)
                else:
                    if vaga.is_livre_quando_nao_especial:  # Apenas se a vaga puder ser livre
                        vagas_livres.append(vaga)
                        print(f"Vaga PNE {vaga.numero} incluída nas vagas livres.")

        # **Sorteio para PNE sem restrição (vagas duplas PNE)**
        random.shuffle(vagas_pne_duplas)  # Aleatoriza as vagas PNE duplas
        for vaga in vagas_pne_duplas:
            # Filtrar apartamentos que estão no mesmo subsolo da vaga
            apartamentos_elegiveis = [ap for ap in apartamentos_pne_sem_restricao if ap.subsolo == vaga.subsolo]
            if apartamentos_elegiveis:
                apartamento_escolhido = random.choice(apartamentos_elegiveis)
                Sorteio.objects.create(apartamento=apartamento_escolhido, vaga=vaga)
                print(f"Sorteado (PNE - sem restrição): {apartamento_escolhido.numero} para a vaga PNE Dupla {vaga.numero}")
                apartamentos_pne_sem_restricao.remove(apartamento_escolhido)
                apartamentos_com_vaga.append(apartamento_escolhido)
            else:
                if vaga.is_livre_quando_nao_especial:  # Apenas se a vaga puder ser livre
                    vagas_livres.append(vaga)
                    print(f"Vaga PNE Dupla {vaga.numero} incluída nas vagas livres.")

        # **Sorteio Aleatório para Idosos (apenas vagas livres)**
        random.shuffle(vagas_idosos)  # Aleatoriza as vagas Idosos livres
        for vaga in vagas_idosos:
            apartamentos_elegiveis = [ap for ap in apartamentos_idoso if ap.subsolo == vaga.subsolo and not ap.apenas_dupla]
            if apartamentos_elegiveis:
                apartamento_escolhido = random.choice(apartamentos_elegiveis)
                Sorteio.objects.create(apartamento=apartamento_escolhido, vaga=vaga)
                print(f"Sorteado: {apartamento_escolhido.numero} para a vaga Idoso {vaga.numero}")
                apartamentos_idoso.remove(apartamento_escolhido)
                apartamentos_com_vaga.append(apartamento_escolhido)
            else:
                if vaga.is_livre_quando_nao_especial:
                    vagas_livres.append(vaga)

        # **Sorteio Aleatório para Idosos (apenas vagas duplas)**
        random.shuffle(vagas_idosos_duplas)
        for vaga in vagas_idosos_duplas:
            apartamentos_elegiveis = [ap for ap in apartamentos_idoso if ap.subsolo == vaga.subsolo and not ap.apenas_livre]
            if apartamentos_elegiveis:
                apartamento_escolhido = random.choice(apartamentos_elegiveis)
                Sorteio.objects.create(apartamento=apartamento_escolhido, vaga=vaga)
                apartamentos_idoso.remove(apartamento_escolhido)
                apartamentos_com_vaga.append(apartamento_escolhido)
            else:
                if vaga.is_livre_quando_nao_especial:
                    vagas_livres.append(vaga)

        # **Sorteio de apartamentos restantes para vagas livres**
        apartamentos_disponiveis = apartamentos_normais + [ap for ap in apartamentos_pne_apenas_livre + apartamentos_pne_sem_restricao + apartamentos_idoso if ap not in apartamentos_com_vaga]

        random.shuffle(apartamentos_disponiveis)
        for apartamento in apartamentos_disponiveis:
            vagas_filtradas = [v for v in vagas_livres if v.subsolo == apartamento.subsolo]
            if vagas_filtradas:
                vaga_escolhida = random.choice(vagas_filtradas)
                Sorteio.objects.create(apartamento=apartamento, vaga=vaga_escolhida)
                print(f"Sorteado: {apartamento.numero} para a vaga Livre {vaga_escolhida.numero}")
                vagas_livres.remove(vaga_escolhida)
            else:
                print(f"Sem vagas disponíveis para o apartamento {apartamento.numero} no subsolo {apartamento.subsolo}")

        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        return redirect('tres_coelhos_sorteio')

    else:
        sorteio_iniciado = request.session.get('sorteio_iniciado', False)
        resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas = resultados_sorteio.exists()

        return render(request, 'tres_coelhos/tres_coelhos_sorteio.html', {
            'resultados_sorteio': resultados_sorteio,
            'vagas_atribuidas': vagas_atribuidas,
            'sorteio_iniciado': sorteio_iniciado,
            'horario_conclusao': request.session.get('horario_conclusao', '')
        })



def tres_coelhos_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteio_novo.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['B8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'B{linha}'] = sorteio.apartamento.numero  # Número do apartamento
        ws[f'C{linha}'] = sorteio.vaga.numero  # Número da vaga
        ws[f'D{linha}'] = f'Subsolo {sorteio.vaga.subsolo}'  # Subsolo
        ws[f'E{linha}'] = sorteio.vaga.get_tipo_display()  # Tipo da vaga
        ws[f'F{linha}'] = sorteio.vaga.get_especial_display()  # Especialidade da vaga
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resultado_sorteio_tres_coelhos.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response


# def tres_coelhos_qrcode(request):
#     # Obter todos os apartamentos para preencher o dropdown
#     apartamentos_disponiveis = Apartamento.objects.all()
    
#     # Obter o apartamento selecionado através do filtro (via GET)
#     numero_apartamento = request.GET.get('apartamento')
    
#     # Inicializar a variável de resultados filtrados como None
#     resultados_filtrados = []

#     # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
#     if numero_apartamento:
#         # Buscar os sorteios para vagas livres
#         sorteios_livres = Sorteio.objects.filter(apartamento__numero=numero_apartamento)
        
#         # Buscar os sorteios para vagas duplas
#         sorteios_duplas = SorteioDupla.objects.filter(apartamento__numero=numero_apartamento)

#         # Unir os resultados de vagas livres e vagas duplas
#         resultados_filtrados = list(sorteios_livres) + list(sorteios_duplas)

#     return render(request, 'tres_coelhos/tres_coelhos_qrcode.html', {
#         'resultados_filtrados': resultados_filtrados,
#         'apartamento_selecionado': numero_apartamento,
#         'apartamentos_disponiveis': apartamentos_disponiveis,
#     })


def tres_coelhos_qrcode(request):
    # Obter todos os apartamentos para preencher o dropdown
    apartamentos_disponiveis = Apartamento.objects.all()
    
    # Obter o apartamento selecionado através do filtro (via GET)
    numero_apartamento = request.GET.get('apartamento')
    
    # Inicializar a variável de resultados filtrados como uma lista vazia
    resultados_filtrados = []

    # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
    if numero_apartamento:
        # Buscar os sorteios para o apartamento selecionado
        sorteios_duplas = SorteioDupla.objects.filter(apartamento__numero=numero_apartamento)

        # Armazenar os resultados filtrados
        resultados_filtrados = sorteios_duplas

    return render(request, 'tres_coelhos/tres_coelhos_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })



def tres_coelhos_dupla(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio de duplas
        SorteioDupla.objects.all().delete()
        print("Sorteios anteriores apagados (duplas).")

        # Obter todas as duplas de apartamentos (pré-selecionadas)
        duplas_apartamentos = list(DuplaApartamentos.objects.all())
        
        # Capturar IDs de apartamentos que fazem parte das duplas
        apartamentos_em_duplas_ids = [dupla.apartamento_1.id for dupla in duplas_apartamentos] + \
                                    [dupla.apartamento_2.id for dupla in duplas_apartamentos if dupla.apartamento_2]

        # Obter apartamentos que não foram sorteados e que não fazem parte de nenhuma dupla
        apartamentos_nao_sorteados = list(Apartamento.objects.exclude(id__in=apartamentos_em_duplas_ids).exclude(sorteio__isnull=False))

        print(f"Duplas formadas: {len(duplas_apartamentos)}")
        print(f"Apartamentos não sorteados: {len(apartamentos_nao_sorteados)}")

        # Obter as vagas duplas disponíveis (para duplas) filtrando por subsolo
        vagas_duplas = list(Vaga.objects.filter(tipo='DUPLA', sorteio__isnull=True))

        print(f"Vagas duplas disponíveis: {len(vagas_duplas)}")

        # **Sorteio de vagas duplas para apartamentos em duplas**
        random.shuffle(duplas_apartamentos)
        for dupla in duplas_apartamentos:
            # Filtrar as vagas de acordo com o subsolo do primeiro apartamento da dupla
            vagas_duplas_subsolo = [vaga for vaga in vagas_duplas if vaga.subsolo == dupla.apartamento_1.subsolo]
            
            if vagas_duplas_subsolo:
                vaga_escolhida_1 = random.choice(vagas_duplas_subsolo)
                vaga_escolhida_2 = vaga_escolhida_1.dupla_com  # Vaga dupla associada

                # Criar o sorteio da dupla
                SorteioDupla.objects.create(apartamento=dupla.apartamento_1, vaga=vaga_escolhida_1)
                SorteioDupla.objects.create(apartamento=dupla.apartamento_2, vaga=vaga_escolhida_2)

                print(f"Sorteado: Apartamento {dupla.apartamento_1.numero} -> Vaga {vaga_escolhida_1.numero}")
                print(f"Sorteado: Apartamento {dupla.apartamento_2.numero} -> Vaga {vaga_escolhida_2.numero}")

                # Remover as vagas duplas da lista de disponíveis
                vagas_duplas.remove(vaga_escolhida_1)
            else:
                print(f"Sem vagas duplas disponíveis no subsolo {dupla.apartamento_1.subsolo}")
                break  # Sem mais vagas duplas no subsolo

        # **Sorteio para apartamentos não sorteados**
        # Apartamentos que não formaram duplas vão concorrer às vagas duplas restantes
        random.shuffle(apartamentos_nao_sorteados)
        
        for apartamento in apartamentos_nao_sorteados:
            # Filtrar as vagas de acordo com o subsolo do apartamento
            vagas_restantes_subsolo = [vaga for vaga in vagas_duplas if vaga.subsolo == apartamento.subsolo]

            if vagas_restantes_subsolo:
                vaga_escolhida = random.choice(vagas_restantes_subsolo)
                SorteioDupla.objects.create(apartamento=apartamento, vaga=vaga_escolhida)
                print(f"Sorteado: Apartamento {apartamento.numero} -> Vaga {vaga_escolhida.numero}")
                vagas_duplas.remove(vaga_escolhida)
            else:
                print(f"Sem vagas disponíveis no subsolo {apartamento.subsolo}")
                break  # Sem mais vagas disponíveis no subsolo

        # Armazenar informações do sorteio na sessão
        request.session['sorteio_dupla_iniciado'] = True
        request.session['horario_conclusao_dupla'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        return redirect('tres_coelhos_dupla')

    else:
        sorteio_iniciado = request.session.get('sorteio_dupla_iniciado', False)
        resultados_sorteio = SorteioDupla.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas = resultados_sorteio.exists()

        return render(request, 'tres_coelhos/tres_coelhos_dupla.html', {
            'resultados_sorteio': resultados_sorteio,
            'vagas_atribuidas': vagas_atribuidas,
            'sorteio_iniciado': sorteio_iniciado,
            'horario_conclusao': request.session.get('horario_conclusao_dupla', '')
        })



def tres_coelhos_dupla_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteio_novo2.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio = SorteioDupla.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'A{linha}'] = sorteio.apartamento.numero  # Número do apartamento
        ws[f'B{linha}'] = sorteio.vaga.numero  # Número da vaga
        ws[f'C{linha}'] = f'Subsolo {sorteio.vaga.subsolo}'  # Subsolo
        ws[f'D{linha}'] = sorteio.vaga.get_tipo_display()  # Tipo da vaga
        ws[f'E{linha}'] = sorteio.vaga.get_especial_display()  # Especialidade da vaga

        # Adicionar "Dupla Com" (número da vaga associada, se houver)
        ws[f'F{linha}'] = sorteio.vaga.dupla_com.numero if sorteio.vaga.dupla_com else "N/A"  # Número da vaga dupla

        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resultado_sorteio_dupla_tres_coelhos.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response


def tres_coelhos_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        SorteioDupla.objects.all().delete()
        return redirect('tres_coelhos_sorteio')
    else:
        return render(request, 'tres_coelhos/tres_coelhos_zerar.html')
    

def tres_coelhos_resultado(request):
    # Obtenha os resultados do sorteio e do sorteio duplo
    resultados_sorteio = list(Sorteio.objects.all())
    resultados_sorteio_dupla = list(SorteioDupla.objects.all())

    # Combine os resultados em uma única lista
    resultados_combinados = resultados_sorteio + resultados_sorteio_dupla

    # Ordene os resultados combinados pelo ID
    resultados_combinados.sort(key=lambda x: x.id)

    # Crie uma lista unificada com dados consistentes
    resultados_formatados = []
    for sorteio in resultados_combinados:
        apartamento_numero = sorteio.apartamento.numero if hasattr(sorteio, 'apartamento') else '-'
        vaga_numero = sorteio.vaga.numero if hasattr(sorteio, 'vaga') else '-'
        vaga_subsolo = sorteio.vaga.subsolo if hasattr(sorteio, 'vaga') else '-'
        tipo_vaga = sorteio.vaga.get_tipo_display() if hasattr(sorteio, 'vaga') else '-'
        especialidade = sorteio.vaga.get_especial_display() if hasattr(sorteio, 'vaga') else '-'
        dupla_com_numero = sorteio.vaga.dupla_com.numero if hasattr(sorteio.vaga, 'dupla_com') and sorteio.vaga.dupla_com else '-'

        resultados_formatados.append({
            'apartamento': apartamento_numero,
            'vaga': vaga_numero,
            'subsolo': vaga_subsolo,
            'tipo_vaga': tipo_vaga,
            'especialidade': especialidade,
            'dupla_com': dupla_com_numero
        })

    # Passe os dados formatados para o contexto
    context = {
        'resultados_combinados': resultados_formatados
    }

    return render(request, 'tres_coelhos/tres_coelhos_resultado.html', context)
