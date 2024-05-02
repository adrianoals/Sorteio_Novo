from django.shortcuts import render, redirect
from .models import Apartamento, Vaga, Sorteio, ApartamentoMoto, VagaMoto, SorteioMoto
from django.utils import timezone
import random
from django.contrib.admin.views.decorators import staff_member_required

# Excel
from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone


# @staff_member_required
# def lyon(request):
#     if request.method == 'POST':
#         # Limpar registros anteriores de sorteio
#         Sorteio.objects.all().delete()
        
#         # Obter todos os apartamentos e grupos de vagas
#         apartamentos = list(Apartamento.objects.all())
#         vagas = list(Vaga.objects.all())

#         # Certifique-se de que existem vagas suficientes para todos os apartamentos
#         if len(vagas) >= len(apartamentos):
#             random.shuffle(vagas)

#             for apartamento in apartamentos:
#                 vaga_selecionada = vagas.pop()
#                 Sorteio.objects.create(
#                     apartamento=apartamento, 
#                     vaga=vaga_selecionada
#                 )
#         else:
           
#             pass
        
#         # Armazenar informações do sorteio na sessão
#         request.session['sorteio_iniciado_nc'] = True
#         request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

#         return redirect('lyon')
    
#     else:
#         sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
#         resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
#         vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

#         return render(request, 'lyon/lyon.html', {
#             'resultados_sorteio_nc': resultados_sorteio_nc,
#             'vagas_atribuidas_nc': vagas_atribuidas_nc,
#             'sorteio_iniciado_nc': sorteio_iniciado_nc,
#             'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
#         })

from django.db import transaction
from django.shortcuts import redirect, render
from .models import Apartamento, Vaga, Sorteio
from django.contrib.admin.views.decorators import staff_member_required
import random
from django.utils import timezone

@staff_member_required
def lyon(request):
    if request.method == 'POST':
        with transaction.atomic():  # Usar transação para garantir atomicidade
            # Limpar registros anteriores de sorteio
            Sorteio.objects.all().delete()

            # Obter todos os apartamentos e vagas
            apartamentos = list(Apartamento.objects.all())
            vagas = list(Vaga.objects.all())

            # Exceção 1 - PCD
            apartamentos_pcd = [ap for ap in apartamentos if ap.pcd]
            vagas_pcd = [vaga for vaga in vagas if vaga.pcd]
            for ap in apartamentos_pcd:
                if vagas_pcd:
                    vaga_selecionada = random.choice(vagas_pcd)
                    vagas_pcd.remove(vaga_selecionada)
                    vagas.remove(vaga_selecionada)
                    Sorteio.objects.create(apartamento=ap, vaga=vaga_selecionada)

            # Exceção 2 - Vagas proibidas
            vagas_proibidas = ["T-19/19A", "T-20/20A", "S1-46/46A", "S1-47/47A", "S2-49/49", "S2-64/64A", "S2-70/70A", "S2-71/71A", "S2-76/76A"]
            apartamentos_restritos = ["54", "82", "92", "103", "111", "121", "193"]

            # Exceção 3 - Vagas cobertas
            apartamentos_cobertos = ['14', '32', '41', '42', '43', '93', '104', '131', '134', '163', '164', '173', '174']
            vagas_cobertas = [vaga for vaga in vagas if vaga.coberta]

            # Exceção 4 - Vagas cobertas e restritas
            apartamentos_cobertos_restritos = ['22', '192']
            vagas_cobertas_restritas = [vaga for vaga in vagas_cobertas if vaga.vaga not in vagas_proibidas]

            # Sorteio para apartamentos que necessitam de vagas cobertas restritas
            for numero in apartamentos_cobertos_restritos:
                ap = next((a for a in apartamentos if a.numero_apartamento == numero), None)
                if ap:
                    vaga_selecionada = random.choice(vagas_cobertas_restritas)
                    vagas_cobertas.remove(vaga_selecionada)
                    vagas.remove(vaga_selecionada)
                    Sorteio.objects.create(apartamento=ap, vaga=vaga_selecionada)

            # Sorteio para apartamentos que necessitam de vagas cobertas
            for numero in apartamentos_cobertos:
                ap = next((a for a in apartamentos if a.numero_apartamento == numero), None)
                if ap:
                    vaga_selecionada = random.choice(vagas_cobertas)
                    vagas_cobertas.remove(vaga_selecionada)
                    vagas.remove(vaga_selecionada)
                    Sorteio.objects.create(apartamento=ap, vaga=vaga_selecionada)

            # Sorteio aleatório dos restantes considerando restrições
            for ap in apartamentos:
                if ap not in [s.apartamento for s in Sorteio.objects.all()]:
                    vagas_disponiveis = [v for v in vagas if (ap.numero_apartamento not in apartamentos_restritos or v.vaga not in vagas_proibidas)]
                    if vagas_disponiveis:
                        vaga_selecionada = random.choice(vagas_disponiveis)
                        vagas.remove(vaga_selecionada)
                        Sorteio.objects.create(apartamento=ap, vaga=vaga_selecionada)

            # Armazenar informações do sorteio na sessão
            request.session['sorteio_iniciado_nc'] = True
            request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

            return redirect('lyon')

    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()

        return render(request, 'lyon/lyon.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })




@staff_member_required
def lyon_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('lyon')
    else:
        return render(request, 'lyon/lyon_zerar.html')


def lyon_excel(request):
    caminho_modelo = 'static/assets/modelos/sorteio_lyon.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'E{linha}'] = sorteio.vaga.vaga
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio_lyon.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response


def lyon_qrcode(request):
    apartamentos_disponiveis = Apartamento.objects.all()  # Adiciona esta linha
    numero_apartamento = request.GET.get('apartamento')
    resultados_filtrados = None
    if numero_apartamento:
        resultados_filtrados = Sorteio.objects.filter(apartamento__numero_apartamento=numero_apartamento)
    
    return render(request, 'lyon/lyon_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,  # Certifique-se de adicionar esta linha
    })



@staff_member_required
def lyon_moto(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        SorteioMoto.objects.all().delete()
        
        # Obter todos os apartamentos e grupos de vagas
        apartamentos = list(ApartamentoMoto.objects.all())
        vagas = list(VagaMoto.objects.all())

        # Certifique-se de que existem vagas suficientes para todos os apartamentos
        if len(vagas) >= len(apartamentos):
            random.shuffle(vagas)

            for apartamento in apartamentos:
                vaga_selecionada = vagas.pop()
                SorteioMoto.objects.create(
                    apartamento=apartamento, 
                    vaga=vaga_selecionada
                )
        else:
           
            pass
        
        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('lyon_moto')
    
    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = SorteioMoto.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'lyon/lyon_moto.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })
    

@staff_member_required
def lyon_moto_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('lyon_moto')
    else:
        return render(request, 'lyon/lyon_moto_zerar.html')


def lyon_moto_excel(request):
    caminho_modelo = 'static/assets/modelos/sorteio_lyon.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'E{linha}'] = sorteio.vaga.vaga
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio_lyon.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response


def lyon_moto_qrcode(request):
    apartamentos_disponiveis = ApartamentoMoto.objects.all()  # Adiciona esta linha
    numero_apartamento = request.GET.get('apartamento')
    resultados_filtrados = None
    if numero_apartamento:
        resultados_filtrados = Sorteio.objects.filter(apartamento__numero_apartamento=numero_apartamento)
    
    return render(request, 'lyon/lyon_moto_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,  # Certifique-se de adicionar esta linha
    })


