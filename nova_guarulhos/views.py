from django.shortcuts import render, redirect
from .models import Apartamento, Vaga, Sorteio
from django.utils import timezone
import random
from django.contrib.admin.views.decorators import staff_member_required

# Excel
from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages


@staff_member_required
def ng_aleatorio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()
        
        # Obter todos os apartamentos e grupos de vagas
        apartamentos = list(Apartamento.objects.all())
        vagas = list(Vaga.objects.all())

        # Certifique-se de que existem vagas suficientes para todos os apartamentos
        if len(vagas) >= len(apartamentos):
            random.shuffle(vagas)

            for apartamento in apartamentos:
                vaga_selecionada = vagas.pop()
                Sorteio.objects.create(
                    apartamento=apartamento, 
                    vaga=vaga_selecionada
                )
        else:
           
            pass
        
        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('ng_aleatorio')
    
    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'nova_guarulhos/ng_aleatorio.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })


@staff_member_required
def ng_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('ng_aleatorio')
    else:
        return render(request, 'nova_guarulhos/ng_zerar.html')


# def ng_excel(request):
#     caminho_modelo = 'static/assets/modelos/sorteio_porcelana.xlsx'

#     wb = load_workbook(caminho_modelo)
#     ws = wb.active

#     # Ordenar os resultados do sorteio pelo ID do apartamento
#     resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

#     horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
#     ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

#     linha = 10
#     for sorteio in resultados_sorteio_nc:
#         ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
#         ws[f'E{linha}'] = sorteio.vaga.vaga
#         linha += 1

#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     nome_arquivo = "resultado_sorteio.xlsx"
#     response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

#     wb.save(response)

#     return response


# def ng_qrcode(request):
#     apartamentos_disponiveis = Apartamento.objects.all()  # Adiciona esta linha
#     numero_apartamento = request.GET.get('apartamento')
#     resultados_filtrados = None
#     if numero_apartamento:
#         resultados_filtrados = Sorteio.objects.filter(apartamento__numero_apartamento=numero_apartamento)
    
#     return render(request, 'porcelana/porcelana_qrcode.html', {
#         'resultados_filtrados': resultados_filtrados,
#         'apartamento_selecionado': numero_apartamento,
#         'apartamentos_disponiveis': apartamentos_disponiveis,  # Certifique-se de adicionar esta linha
#     })


@staff_member_required
def ng_inicio(request):
    	return render(request, 'nova_guarulhos/ng_inicio.html')


@staff_member_required
def ng_presenca(request):
    if request.method == 'POST':
        apartamentos = Apartamento.objects.all()
        for item in apartamentos:
            item.presenca = request.POST.get('presenca' + str(item.id)) == 'True'
            item.pcd = request.POST.get('pcd' + str(item.id)) == 'True'
            item.idoso = request.POST.get('idoso' + str(item.id)) == 'True'
            item.adimplentes = request.POST.get('adimplentes' + str(item.id)) == 'True'
            item.save()
        return redirect('ng_filtrar')  # Redireciona para a rota 'ng_filtrar'
    apartamentos = Apartamento.objects.all()
    return render(request, 'nova_guarulhos/ng_presenca.html', {"lista_de_presenca": apartamentos})

@staff_member_required
def ng_filtrar(request):
    lista_de_presenca = Apartamento.objects.none()  # Inicia com uma queryset vazia
    if request.method == 'POST':
        lista_de_presenca = Apartamento.objects.all()  # Recupera todos os objetos quando o formulário é submetido
        if 'presentes' in request.POST:
            lista_de_presenca = lista_de_presenca.filter(presenca=True)
        if 'ausentes' in request.POST:
            lista_de_presenca = lista_de_presenca.filter(presenca=False)
    return render(request, 'nova_guarulhos/ng_filtrar.html', {"lista_de_presenca": lista_de_presenca})


@staff_member_required
def ng_apartamento(request):
    # Seleciona todos os apartamentos com presença confirmada e que ainda não têm vaga sorteada
    apartamentos_disponiveis = Apartamento.objects.filter(presenca=True).exclude(sorteio__apartamento__isnull=False)
    # Seleciona todas as vagas que ainda não estão em um sorteio
    vagas_disponiveis = Vaga.objects.exclude(sorteio__vaga__isnull=False)

    sorteio_finalizado = not apartamentos_disponiveis.exists() or not vagas_disponiveis.exists()
    item_de_presenca = None  # Inicializa a variável

    if request.method == 'POST':
        if 'realizar_sorteio' in request.POST and apartamentos_disponiveis.exists():
            # Sorteio de um apartamento aleatório
            apartamento_escolhido = random.choice(list(apartamentos_disponiveis))
            messages.success(request, f'Apartamento {apartamento_escolhido.numero_apartamento} foi selecionado para atribuição de vaga. Escolha a vaga agora.')

            # Armazena o apartamento escolhido para permitir seleção de vaga
            item_de_presenca = apartamento_escolhido

            # Renderiza a página com o apartamento sorteado para escolha da vaga
            return render(request, 'nova_guarulhos/ng_apartamento.html', {
                'sorteio_finalizado': sorteio_finalizado,
                'apartamentos_disponiveis': apartamentos_disponiveis,
                'vagas_disponiveis': vagas_disponiveis,
                'item_de_presenca': item_de_presenca
            })

        # Seção para escolher vaga para o apartamento selecionado
        if 'vaga_selecionada' in request.POST and 'apartamento_id' in request.POST:
            apartamento_id = request.POST.get('apartamento_id')
            vaga_selecionada = request.POST.get('vaga_selecionada')
            apartamento = Apartamento.objects.get(id=apartamento_id)
            vaga = Vaga.objects.get(vaga=vaga_selecionada)

            # Cria um novo registro de sorteio associando o apartamento à vaga escolhida
            novo_sorteio = Sorteio(apartamento=apartamento, vaga=vaga)
            novo_sorteio.save()

            messages.success(request, f'Vaga {vaga.vaga} confirmada para o apartamento {apartamento.numero_apartamento} com sucesso!')
            return redirect('ng_apartamento')

    return render(request, 'nova_guarulhos/ng_apartamento.html', {
        'sorteio_finalizado': sorteio_finalizado,
        'apartamentos_disponiveis': apartamentos_disponiveis,
        'vagas_disponiveis': vagas_disponiveis,
        'item_de_presenca': item_de_presenca
    })


@staff_member_required
def ng_pcd(request):
    # Seleciona todos os apartamentos com presença confirmada e que ainda não têm vaga sorteada
    apartamentos_pcd = list(Apartamento.objects.filter(presenca=True, pcd=True).exclude(sorteio__apartamento__isnull=False))
    apartamentos_idoso = list(Apartamento.objects.filter(presenca=True, idoso=True, pcd=False).exclude(sorteio__apartamento__isnull=False))
    # Seleciona todas as vagas que ainda não estão em um sorteio
    vagas_disponiveis = list(Vaga.objects.exclude(sorteio__vaga__isnull=False))

    sorteio_finalizado = not (apartamentos_pcd or apartamentos_idoso) or not vagas_disponiveis
    lista_de_presenca = []

    if request.method == 'POST':
        if 'confirmar_vagas' in request.POST:
            for key, value in request.POST.items():
                if key.startswith('apartamento_'):
                    apartamento_id = key.split('_')[1]
                    apartamento = Apartamento.objects.get(id=apartamento_id)
                    vaga = Vaga.objects.get(vaga=value)
                    # Cria um novo registro de sorteio associando o apartamento à vaga escolhida
                    novo_sorteio = Sorteio(apartamento=apartamento, vaga=vaga)
                    novo_sorteio.save()
            messages.success(request, 'Vagas confirmadas com sucesso!')
            return redirect('ng_pcd')

        if 'realizar_sorteio' in request.POST:
            random.shuffle(apartamentos_pcd)
            random.shuffle(apartamentos_idoso)
            lista_de_presenca = apartamentos_pcd + apartamentos_idoso

            return render(request, 'nova_guarulhos/ng_pcd.html', {
                'sorteio_finalizado': sorteio_finalizado,
                'vagas_disponiveis': vagas_disponiveis,
                'lista_de_presenca': lista_de_presenca
            })

    return render(request, 'nova_guarulhos/ng_pcd.html', {
        'sorteio_finalizado': sorteio_finalizado,
        'vagas_disponiveis': vagas_disponiveis,
        'lista_de_presenca': lista_de_presenca
    })


@staff_member_required
def ng_blocos(request):
    # Filtra os blocos que têm apartamentos presentes, adimplentes e com vagas a serem atribuídas
    blocos_com_apartamentos_validos = list(Apartamento.objects.filter(presenca=True, adimplentes=True).exclude(sorteio__apartamento__isnull=False).values_list('bloco', flat=True).distinct())

    sorteio_realizado = False

    if request.method == 'POST':
        random.shuffle(blocos_com_apartamentos_validos)
        request.session['ordem_blocos'] = blocos_com_apartamentos_validos
        sorteio_realizado = True
        print(f"Ordem dos blocos sorteada e armazenada na sessão: {blocos_com_apartamentos_validos}")

    ordem_blocos = request.session.get('ordem_blocos', [])
    print(f"Ordem dos blocos na sessão ao renderizar: {ordem_blocos}")

    return render(request, 'nova_guarulhos/ng_blocos.html', {
        'ordem_blocos': ordem_blocos,
        'sorteio_realizado': sorteio_realizado
    })


@staff_member_required
def ng_adimplentes(request):
    ordem_blocos = request.session.get('ordem_blocos', [])
    print(f"Ordem dos blocos na sessão ao iniciar ng_adimplentes: {ordem_blocos}")

    apartamentos_adimplentes = []

    # Mantém a ordem dos blocos e sorteia aleatoriamente os apartamentos dentro de cada bloco
    for bloco in ordem_blocos:
        apartamentos_do_bloco = list(Apartamento.objects.filter(presenca=True, adimplentes=True, bloco=bloco).exclude(sorteio__apartamento__isnull=False))
        random.shuffle(apartamentos_do_bloco)
        apartamentos_adimplentes.extend(apartamentos_do_bloco)

    print(f"Apartamentos adimplentes sorteados dentro dos blocos: {[(apt.bloco, apt.numero_apartamento) for apt in apartamentos_adimplentes]}")

    vagas_disponiveis = list(Vaga.objects.exclude(sorteio__vaga__isnull=False))
    sorteio_finalizado = not apartamentos_adimplentes or not vagas_disponiveis
    lista_de_presenca = apartamentos_adimplentes if 'realizar_sorteio' in request.POST else []

    if request.method == 'POST':
        if 'confirmar_vagas' in request.POST:
            for key, value in request.POST.items():
                if key.startswith('apartamento_'):
                    apartamento_id = key.split('_')[1]
                    apartamento = Apartamento.objects.get(id=apartamento_id)
                    vaga = Vaga.objects.get(vaga=value)
                    novo_sorteio = Sorteio(apartamento=apartamento, vaga=vaga)
                    novo_sorteio.save()
            messages.success(request, 'Vagas confirmadas com sucesso!')
            return redirect('ng_adimplentes')

    return render(request, 'nova_guarulhos/ng_adimplentes.html', {
        'sorteio_finalizado': sorteio_finalizado,
        'vagas_disponiveis': vagas_disponiveis,
        'lista_de_presenca': lista_de_presenca,
        'ordem_blocos': ordem_blocos,
    })


def ng_excel(request):
    caminho_modelo = 'static/assets/modelos/sorteio_nova_guarulhos.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__bloco', 'apartamento__numero_apartamento').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'C{linha}'] = sorteio.apartamento.bloco
        ws[f'D{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'E{linha}'] = sorteio.vaga.vaga
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response


@staff_member_required
def ng_final(request):
    if request.method == 'POST':
        # Obter todos os apartamentos que ainda não têm vagas atribuídas
        apartamentos_sem_vaga = list(Apartamento.objects.exclude(sorteio__isnull=False))
        vagas_disponiveis = list(Vaga.objects.exclude(id__in=Sorteio.objects.values_list('vaga_id', flat=True)))

        # Verifica se há vagas suficientes para os apartamentos restantes
        if len(vagas_disponiveis) >= len(apartamentos_sem_vaga):
            random.shuffle(vagas_disponiveis)

            for apartamento in apartamentos_sem_vaga:
                vaga_selecionada = vagas_disponiveis.pop()
                Sorteio.objects.create(
                    apartamento=apartamento, 
                    vaga=vaga_selecionada
                )
        else:
            messages.error(request, 'Não há vagas suficientes para todos os apartamentos restantes.')

        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('ng_final')

    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        todos_apartamentos = Apartamento.objects.count()  # Conta todos os apartamentos registrados
        apartamentos_sorteio = Sorteio.objects.count()  # Conta todos os apartamentos com vagas atribuídas
        vagas_atribuidas_completas = todos_apartamentos == apartamentos_sorteio  # Verifica se todos têm vagas atribuídas

        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__bloco', 'apartamento__numero_apartamento').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'nova_guarulhos/ng_final.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', ''),
            'vagas_atribuidas_completas': vagas_atribuidas_completas  # Adiciona essa variável ao contexto
        })


def ng_qrcode(request):
    apartamentos_disponiveis = Apartamento.objects.all()  # Obtém todos os apartamentos disponíveis
    bloco_apartamento = request.GET.get('apartamento')
    resultados_filtrados = None
    if bloco_apartamento:
        bloco, numero_apartamento = bloco_apartamento.split(" - ")
        resultados_filtrados = Sorteio.objects.filter(apartamento__bloco=bloco, apartamento__numero_apartamento=numero_apartamento)
    
    return render(request, 'nova_guarulhos/ng_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': bloco_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,  # Certifique-se de adicionar esta linha
    })
