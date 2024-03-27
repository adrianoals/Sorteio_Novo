from django.shortcuts import render, redirect, get_object_or_404
from .models import Apartamento, Sorteio, Vaga
import random
from django.utils import timezone
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required


@staff_member_required
def cbs_index(request):
    if request.method == 'POST':
        # Limpar registros anteriores
        Sorteio.objects.all().delete()
        
        # Obter todos os apartamentos
        apartamentos = list(Apartamento.objects.exclude(numero_apartamento='Apto 02', bloco__bloco='B'))

        # Separar as vagas por bloco
        vagas_a = list(Vaga.objects.filter(vaga__startswith='Vaga ', vaga__endswith='Bloco A'))
        vagas_b = list(Vaga.objects.filter(vaga__startswith='Vaga ', vaga__endswith='Bloco B'))

        # Remover Vaga 15 Bloco B para atribuição direta
        vaga_especial = None
        for vaga in vagas_b:
            if vaga.vaga == 'Vaga 15 Bloco B':
                vaga_especial = vaga
                break
        if vaga_especial:
            vagas_b.remove(vaga_especial)
        
        # Atribuir Vaga 15 Bloco B ao Apto 02 do Bloco B diretamente
        apto_especial = Apartamento.objects.get(numero_apartamento='Apto 02', bloco__bloco='B')
        Sorteio.objects.create(apartamento=apto_especial, vaga=vaga_especial)

        # Embaralhar as listas de vagas
        random.shuffle(vagas_a)
        random.shuffle(vagas_b)

        for apartamento in apartamentos:
            if apartamento.bloco.bloco == 'A' and vagas_a:
                vaga_selecionada = vagas_a.pop()
            elif apartamento.bloco.bloco == 'B' and vagas_b:
                vaga_selecionada = vagas_b.pop()
            else:
                # Se não houver vagas disponíveis, continuar para o próximo apartamento
                continue
            Sorteio.objects.create(apartamento=apartamento, vaga=vaga_selecionada)

        # Marcar na sessão que o sorteio foi iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh%Mmin%Ss")

        # Redirecionar para a mesma página para mostrar os resultados do sorteio
        return redirect('cbs_index')
    
    else:
        # Verificar se o sorteio já foi iniciado
        sorteio_iniciado = request.session.get('sorteio_iniciado', False)
        resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').all() if sorteio_iniciado else []

        # Variável para controlar a exibição da mensagem e do botão
        vagas_atribuidas = bool(resultados_sorteio)

        return render(request, 'chacara_bom_sucesso/cbs_index.html', {
            'resultados_sorteio': resultados_sorteio,
            'vagas_atribuidas': vagas_atribuidas,
            'sorteio_iniciado': sorteio_iniciado,
            'horario_conclusao': request.session.get('horario_conclusao', '')
        })



@staff_member_required
def cbs_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('cbs_index')
    else:
        return render(request, 'chacara_bom_sucesso/cbs_zerar.html')


from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone
from .models import Sorteio

def cbs_exportar_para_excel(request):
    # Construir o caminho completo para o arquivo modelo
    caminho_modelo = 'static/assets/modelos/sorteio_novo1.xlsx' 

    # Carregar o workbook do modelo
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Obter os resultados do sorteio
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').all()

    # Recuperar o horário de conclusão do sorteio da sessão
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')

    # Escrever o horário de conclusão na célula C8
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Supondo que você queira começar a inserir os dados a partir da linha 10
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'D{linha}'] = sorteio.apartamento.bloco.bloco
        ws[f'E{linha}'] = sorteio.vaga.vaga
        linha += 1

    # Preparar a resposta para enviar o arquivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    # Salvar o workbook modificado no response
    wb.save(response)

    return response


def cbs_filtro_apartamento(request):
    apartamentos_disponiveis = Apartamento.objects.all()
    apartamento_id = request.GET.get('apartamento')  # Este agora é o ID do apartamento

    resultados_filtrados = None
    apartamento_selecionado = None
    if apartamento_id:
        apartamento_selecionado = get_object_or_404(Apartamento, id=apartamento_id)
        resultados_filtrados = Sorteio.objects.filter(apartamento=apartamento_selecionado)
    
    return render(request, 'chacara_bom_sucesso/cbs_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': apartamento_selecionado,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })
