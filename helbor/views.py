from django.shortcuts import render, redirect
from .models import ApartamentoTorre1, VagaTorre1, SorteioTorre1, ApartamentoTorre2, VagaTorre2, SorteioTorre2
from django.utils import timezone
import random
from django.contrib.admin.views.decorators import staff_member_required

# Excel
from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages


# @staff_member_required
# def helbor_torre1(request):
#     if request.method == 'POST':
#         # Limpar registros anteriores de sorteio
#         SorteioTorre1.objects.all().delete()
        
#         # Obter todos os apartamentos e grupos de vagas
#         apartamentos = list(ApartamentoTorre1.objects.all())
#         vagas = list(VagaTorre1.objects.all())

#         # Certifique-se de que existem vagas suficientes para todos os apartamentos
#         if len(vagas) >= len(apartamentos):
#             random.shuffle(vagas)

#             for apartamento in apartamentos:
#                 vaga_selecionada = vagas.pop()
#                 SorteioTorre1.objects.create(
#                     apartamento=apartamento, 
#                     vaga=vaga_selecionada
#                 )
#         else:
           
#             pass
        
#         # Armazenar informações do sorteio na sessão
#         request.session['sorteio_iniciado_nc'] = True
#         request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

#         return redirect('helbor_torre1')
    
#     else:
#         sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
#         resultados_sorteio_nc = SorteioTorre1.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
#         vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

#         return render(request, 'helbor/helbor_torre1.html', {
#             'resultados_sorteio_nc': resultados_sorteio_nc,
#             'vagas_atribuidas_nc': vagas_atribuidas_nc,
#             'sorteio_iniciado_nc': sorteio_iniciado_nc,
#             'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
#         })

# Só com validacao de pcd
# from django.shortcuts import render, redirect
# from .models import ApartamentoTorre1, VagaTorre1, SorteioTorre1
# from django.utils import timezone
# import random
# from django.contrib.admin.views.decorators import staff_member_required

# @staff_member_required
# def helbor_torre1(request):
#     if request.method == 'POST':
#         # Limpar registros anteriores de sorteio
#         SorteioTorre1.objects.all().delete()
        
#         # Predefinições de apartamentos e suas vagas correspondentes
#         predefinicoes = {
#             'Torre 1 - Apartamento 31': 'Grupo 29 - Vagas (126, 127, 128) - Pavimento: 3',
#             'Torre 1 - Apartamento 54': 'Grupo 54 - Vagas (276, 277, 293) - Pavimento: 2',
#             'Torre 1 - Apartamento 83': 'Grupo 21 - Vagas (10, 110, 111) - Pavimento: 3',
#             'Torre 1 - Apartamento 104': 'Grupo 130 - Vagas (280, 281, 282) - Pavimento: 2',
#             'Torre 1 - Apartamento 114': 'Grupo 76 - Vagas (447, 448, 346) - Pavimento: 1',
#             'Torre 1 - Apartamento 144': 'Grupo 83 - Vagas (461, 462, 353) - Pavimento: 1',
#             'Torre 1 - Apartamento 164': 'Grupo 10 - Vagas (51, 52, 134) - Pavimento: 3',
#             'Torre 1 - Apartamento 204': 'Grupo 84 - Vagas (463, 464, 465) - Pavimento: 1'
#         }

#         # Atribuir as vagas predefinidas
#         for apartamento_num, vaga_desc in predefinicoes.items():
#             apartamento = ApartamentoTorre1.objects.get(numero_apartamento=apartamento_num)
#             vaga = VagaTorre1.objects.get(vaga=vaga_desc)
#             SorteioTorre1.objects.create(apartamento=apartamento, vaga=vaga)

#         # Obter todos os apartamentos e vagas, excluindo os predefinidos
#         apartamentos = list(ApartamentoTorre1.objects.exclude(numero_apartamento__in=predefinicoes.keys()))
#         vagas = list(VagaTorre1.objects.exclude(vaga__in=predefinicoes.values()))

#         # Certifique-se de que existem vagas suficientes para os apartamentos restantes
#         if len(vagas) >= len(apartamentos):
#             random.shuffle(vagas)

#             for apartamento in apartamentos:
#                 vaga_selecionada = vagas.pop()
#                 SorteioTorre1.objects.create(
#                     apartamento=apartamento, 
#                     vaga=vaga_selecionada
#                 )
#         else:
#             # Tratar o caso em que não há vagas suficientes para os apartamentos restantes
#             pass
        
#         # Armazenar informações do sorteio na sessão
#         request.session['sorteio_iniciado_nc'] = True
#         request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

#         return redirect('helbor_torre1')
    
#     else:
#         sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
#         resultados_sorteio_nc = SorteioTorre1.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
#         vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

#         return render(request, 'helbor/helbor_torre1.html', {
#             'resultados_sorteio_nc': resultados_sorteio_nc,
#             'vagas_atribuidas_nc': vagas_atribuidas_nc,
#             'sorteio_iniciado_nc': sorteio_iniciado_nc,
#             'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
#         })


from django.shortcuts import render, redirect
from .models import ApartamentoTorre1, VagaTorre1, SorteioTorre1
from django.utils import timezone
import random
from django.contrib.admin.views.decorators import staff_member_required


# Torre 1
@staff_member_required
def helbor_torre1(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        SorteioTorre1.objects.all().delete()
        
        # Predefinições de apartamentos e suas vagas correspondentes
        predefinicoes = {
            'Torre 1 - Apartamento 31': 'Grupo 29 - Vagas (126, 127, 128) - Pavimento: 3',
            'Torre 1 - Apartamento 54': 'Grupo 54 - Vagas (276, 277, 293) - Pavimento: 2',
            'Torre 1 - Apartamento 83': 'Grupo 21 - Vagas (10, 110, 111) - Pavimento: 3',
            'Torre 1 - Apartamento 104': 'Grupo 130 - Vagas (280, 281, 282) - Pavimento: 2',
            'Torre 1 - Apartamento 114': 'Grupo 76 - Vagas (447, 448, 346) - Pavimento: 1',
            'Torre 1 - Apartamento 144': 'Grupo 83 - Vagas (461, 462, 353) - Pavimento: 1',
            'Torre 1 - Apartamento 164': 'Grupo 10 - Vagas (51, 52, 134) - Pavimento: 3',
            'Torre 1 - Apartamento 204': 'Grupo 84 - Vagas (463, 464, 465) - Pavimento: 1'
        }

        # Atribuir as vagas predefinidas
        for apartamento_num, vaga_desc in predefinicoes.items():
            apartamento = ApartamentoTorre1.objects.get(numero_apartamento=apartamento_num)
            vaga = VagaTorre1.objects.get(vaga=vaga_desc)
            SorteioTorre1.objects.create(apartamento=apartamento, vaga=vaga)

        # Opções limitadas para o apartamento 64
        restricoes_apartamento_64 = [
            'Grupo 74 - Vagas (443, 444, 344) - Pavimento: 1',
            'Grupo 73 - Vagas (441, 442, 343) - Pavimento: 1'
        ]

        apartamento_64 = ApartamentoTorre1.objects.get(numero_apartamento="Torre 1 - Apartamento 64")
        vagas_restritas = VagaTorre1.objects.filter(vaga__in=restricoes_apartamento_64)
        vaga_escolhida_64 = random.choice(vagas_restritas)
        SorteioTorre1.objects.create(apartamento=apartamento_64, vaga=vaga_escolhida_64)

        # Obter todos os apartamentos e vagas, excluindo os predefinidos e as vagas já escolhidas para o apartamento 64
        apartamentos = list(ApartamentoTorre1.objects.exclude(
            numero_apartamento__in=[*predefinicoes.keys(), "Torre 1 - Apartamento 64"]
        ))
        vagas = list(VagaTorre1.objects.exclude(
            vaga__in=[*predefinicoes.values(), vaga_escolhida_64.vaga]
        ))

        # Certifique-se de que existem vagas suficientes para os apartamentos restantes
        if len(vagas) >= len(apartamentos):
            random.shuffle(vagas)

            for apartamento in apartamentos:
                vaga_selecionada = vagas.pop()
                SorteioTorre1.objects.create(
                    apartamento=apartamento, 
                    vaga=vaga_selecionada
                )
        else:
            # Tratar o caso em que não há vagas suficientes para os apartamentos restantes
            pass
        
        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('helbor_torre1')
    
    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = SorteioTorre1.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'helbor/helbor_torre1.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })


@staff_member_required
def helbor_zerar_torre1(request):
    if request.method == 'POST':
        SorteioTorre1.objects.all().delete()
        return redirect('helbor_torre1')
    else:
        return render(request, 'helbor/helbor_zerar_torre1.html')


def helbor_excel_torre1(request):
    caminho_modelo = 'static/assets/modelos/sorteio_helbor_torre1.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = SorteioTorre1.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'E{linha}'] = sorteio.vaga.vaga
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio_torre1.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response


def helbor_qrcode_torre1(request):
    apartamentos_disponiveis = ApartamentoTorre1.objects.all()  # Adiciona esta linha
    numero_apartamento = request.GET.get('apartamento')
    resultados_filtrados = None
    if numero_apartamento:
        resultados_filtrados = SorteioTorre1.objects.filter(apartamento__numero_apartamento=numero_apartamento)
    
    return render(request, 'helbor/helbor_qrcode_torre1.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,  # Certifique-se de adicionar esta linha
    })


# Torre 2
@staff_member_required
def helbor_torre2(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        SorteioTorre2.objects.all().delete()
        
        # Predefinições de apartamentos e suas vagas correspondentes
        predefinicoes = {
            'Torre 2 - Apartamento 33': 'Grupo 120 - Vagas (226, 227, 300) - Pavimento: 2',
            'Torre 2 - Apartamento 63': 'Grupo 123 - Vagas (232, 233, 306) - Pavimento: 2',
            'Torre 2 - Apartamento 72': 'Grupo 144 - Vagas (397, 398, 395) - Pavimento: 1',
            'Torre 2 - Apartamento 102': 'Grupo 106 - Vagas (135, 143, 144) - Pavimento: 3',
            'Torre 2 - Apartamento 113': 'Grupo 110 - Vagas (130, 168, 169) - Pavimento: 3',
            'Torre 2 - Apartamento 114': 'Grupo 119 - Vagas (224, 225, 301) - Pavimento: 2',
            'Torre 2 - Apartamento 172': 'Grupo 122 - Vagas (230, 231, 305) - Pavimento: 2',
            'Torre 2 - Apartamento 212': 'Grupo 118 - Vagas (302, 303, 223) - Pavimento: 2',
            'Torre 2 - Apartamento 221': 'Grupo 140 - Vagas (333, 334, 316) - Pavimento: 2',
            'Torre 2 - Apartamento 223': 'Grupo 113 - Vagas (133, 174, 175) - Pavimento: 3',
            'Torre 2 - Apartamento 224': 'Grupo 168 - Vagas (505, 506, 472) - Pavimento: 1'

            }

        # Atribuir as vagas predefinidas
        for apartamento_num, vaga_desc in predefinicoes.items():
            apartamento = ApartamentoTorre2.objects.get(numero_apartamento=apartamento_num)
            vaga = VagaTorre2.objects.get(vaga=vaga_desc)
            SorteioTorre2.objects.create(apartamento=apartamento, vaga=vaga)

        # Obter todos os apartamentos e vagas, excluindo os predefinidos
        apartamentos = list(ApartamentoTorre2.objects.exclude(numero_apartamento__in=predefinicoes.keys()))
        vagas = list(VagaTorre2.objects.exclude(vaga__in=predefinicoes.values()))

        # Certifique-se de que existem vagas suficientes para os apartamentos restantes
        if len(vagas) >= len(apartamentos):
            random.shuffle(vagas)

            for apartamento in apartamentos:
                vaga_selecionada = vagas.pop()
                SorteioTorre2.objects.create(
                    apartamento=apartamento, 
                    vaga=vaga_selecionada
                )
        else:
            # Tratar o caso em que não há vagas suficientes para os apartamentos restantes
            pass
        
        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('helbor_torre2')
    
    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = SorteioTorre2.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'helbor/helbor_torre2.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })



@staff_member_required
def helbor_zerar_torre2(request):
    if request.method == 'POST':
        SorteioTorre2.objects.all().delete()
        return redirect('helbor_torre2')
    else:
        return render(request, 'helbor/helbor_zerar_torre2.html')


def helbor_excel_torre2(request):
    caminho_modelo = 'static/assets/modelos/sorteio_helbor_torre2.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = SorteioTorre2.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'E{linha}'] = sorteio.vaga.vaga
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio_torre2.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response


def helbor_qrcode_torre2(request):
    apartamentos_disponiveis = ApartamentoTorre2.objects.all()  # Adiciona esta linha
    numero_apartamento = request.GET.get('apartamento')
    resultados_filtrados = None
    if numero_apartamento:
        resultados_filtrados = SorteioTorre2.objects.filter(apartamento__numero_apartamento=numero_apartamento)
    
    return render(request, 'helbor/helbor_qrcode_torre2.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,  # Certifique-se de adicionar esta linha
    })

