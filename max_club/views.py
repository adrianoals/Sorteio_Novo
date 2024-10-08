from django.shortcuts import render, redirect
from .models import Apartamento, Vaga, Sorteio, ApartamentoBike, VagaBike, SorteioBike
from django.utils import timezone
import random
from django.contrib.admin.views.decorators import staff_member_required

# Excel
from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone
from .models import Sorteio, SorteioBike

    
# @staff_member_required
# def max_club(request):
#     if request.method == 'POST':
#         # Limpar registros anteriores de sorteio
#         Sorteio.objects.all().delete()
        
#         # Obter todos os apartamentos e grupos de vagas
#         apartamentos = list(Apartamento.objects.all())
#         vagas = list(Vaga.objects.all())
        
#         # Definir a vaga especial
#         vaga_especial = Vaga.objects.get(vaga="Vaga PNE Subsolo: 01")
        
#         # Sortear a vaga especial para o apartamento 904
#         apartamento_especial = None
#         for apt in apartamentos:
#             if apt.numero_apartamento == "904":
#                 apartamento_especial = apt
#                 break
        
#         if apartamento_especial:
#             # Criar sorteio para o apartamento especial
#             Sorteio.objects.create(
#                 apartamento=apartamento_especial,
#                 vaga=vaga_especial
#             )
#             # Remover apartamento especial e a vaga especial das listas
#             apartamentos.remove(apartamento_especial)
#             vagas.remove(vaga_especial)
        
#         # Certifique-se de que existem vagas suficientes para os outros apartamentos
#         if len(vagas) >= len(apartamentos):
#             random.shuffle(vagas)
#             for apartamento in apartamentos:
#                 vaga_selecionada = vagas.pop()
#                 Sorteio.objects.create(
#                     apartamento=apartamento, 
#                     vaga=vaga_selecionada
#                 )
#         else:
#             pass
        
#         # Armazenar informações do sorteio na sessão
#         request.session['sorteio_iniciado_nc'] = True
#         request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

#         return redirect('max_club')
    
#     else:
#         sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
#         resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
#         vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

#         return render(request, 'max_club/max_club.html', {
#             'resultados_sorteio_nc': resultados_sorteio_nc,
#             'vagas_atribuidas_nc': vagas_atribuidas_nc,
#             'sorteio_iniciado_nc': sorteio_iniciado_nc,
#             'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
#         })



# Aqui funciona
# from django.shortcuts import render, redirect
# from .models import Apartamento, Vaga, Sorteio
# from django.utils import timezone
# import random
# from django.contrib.admin.views.decorators import staff_member_required

# @staff_member_required
# def max_club(request):
#     if request.method == 'POST':
#         # Limpar registros anteriores de sorteio
#         Sorteio.objects.all().delete()
        
#         # Obter todos os apartamentos e grupos de vagas
#         apartamentos = list(Apartamento.objects.all())
#         vagas = list(Vaga.objects.all())
        
#         # Definir as vagas especiais
#         vaga_pne = Vaga.objects.get(vaga="Vaga PNE Subsolo: 01")
#         vagas_prioritarias = [
#             Vaga.objects.get(vaga="Vaga 63 Idoso Subsolo: 01"),
#             Vaga.objects.get(vaga="Vaga 48 Idoso Subsolo: 01"),
#             Vaga.objects.get(vaga="Vaga 33 Idoso Subsolo: 02"),
#             Vaga.objects.get(vaga="Vaga 32 Idoso Subsolo: 02"),
#             Vaga.objects.get(vaga="Vaga 17 Idoso Subsolo: 02")
#         ]
        
#         # Sortear a vaga PNE entre os apartamentos 904 e 908
#         apartamentos_especiais = [apt for apt in apartamentos if apt.numero_apartamento in ["904", "908"]]
#         apartamento_pne = random.choice(apartamentos_especiais)
#         apartamentos_especiais.remove(apartamento_pne)
#         apartamento_prioritario = apartamentos_especiais[0]

#         # Atribuir a vaga PNE
#         Sorteio.objects.create(
#             apartamento=apartamento_pne,
#             vaga=vaga_pne
#         )
#         apartamentos.remove(apartamento_pne)
#         vagas.remove(vaga_pne)

#         # Atribuir uma das vagas prioritárias ao outro apartamento especial
#         vaga_prioritaria_escolhida = random.choice(vagas_prioritarias)
#         Sorteio.objects.create(
#             apartamento=apartamento_prioritario,
#             vaga=vaga_prioritaria_escolhida
#         )
#         apartamentos.remove(apartamento_prioritario)
#         vagas.remove(vaga_prioritaria_escolhida)
#         vagas_prioritarias.remove(vaga_prioritaria_escolhida)

#         # Adicionar as vagas prioritárias restantes à lista de vagas disponíveis
#         vagas_restantes = vagas_prioritarias + [vaga for vaga in vagas if vaga not in vagas_prioritarias]

#         if len(vagas_restantes) >= len(apartamentos):
#             random.shuffle(vagas_restantes)
#             for apartamento in apartamentos:
#                 vaga_selecionada = vagas_restantes.pop()
#                 Sorteio.objects.create(
#                     apartamento=apartamento, 
#                     vaga=vaga_selecionada
#                 )
        
#         # Armazenar informações do sorteio na sessão
#         request.session['sorteio_iniciado_nc'] = True
#         request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

#         return redirect('max_club')
    
#     else:
#         sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
#         resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
#         vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

#         return render(request, 'max_club/max_club.html', {
#             'resultados_sorteio_nc': resultados_sorteio_nc,
#             'vagas_atribuidas_nc': vagas_atribuidas_nc,
#             'sorteio_iniciado_nc': sorteio_iniciado_nc,
#             'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
#         })


# @staff_member_required
def max_club(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()
        
        # Obter todos os apartamentos e grupos de vagas
        apartamentos = list(Apartamento.objects.all())
        vagas = list(Vaga.objects.all())
        
        # Definir as vagas especiais
        vaga_pne = Vaga.objects.get(vaga="Vaga PNE Subsolo: 01")
        vagas_prioritarias = [
            Vaga.objects.get(vaga="Vaga 63 Idoso Subsolo: 01"),
            Vaga.objects.get(vaga="Vaga 48 Idoso Subsolo: 01"),
            Vaga.objects.get(vaga="Vaga 33 Idoso Subsolo: 02"),
            Vaga.objects.get(vaga="Vaga 32 Idoso Subsolo: 02"),
            Vaga.objects.get(vaga="Vaga 17 Idoso Subsolo: 02")
        ]
        
        # Lista de apartamentos especiais
        apartamentos_especiais = [apt for apt in apartamentos if apt.numero_apartamento in ["904", "908"]]
        
        # Sortear a vaga PNE para um dos apartamentos especiais
        apartamento_pne = random.choice(apartamentos_especiais)
        apartamentos_especiais.remove(apartamento_pne)

        # Atribuir a vaga PNE
        Sorteio.objects.create(
            apartamento=apartamento_pne,
            vaga=vaga_pne
        )
        apartamentos.remove(apartamento_pne)
        vagas.remove(vaga_pne)

        # Atribuir vagas prioritárias aos demais apartamentos especiais
        for apt in apartamentos_especiais:
            if vagas_prioritarias:
                vaga_prioritaria_escolhida = random.choice(vagas_prioritarias)
                Sorteio.objects.create(
                    apartamento=apt,
                    vaga=vaga_prioritaria_escolhida
                )
                apartamentos.remove(apt)
                vagas.remove(vaga_prioritaria_escolhida)
                vagas_prioritarias.remove(vaga_prioritaria_escolhida)

        # Adicionar as vagas prioritárias restantes à lista de vagas disponíveis
        vagas_restantes = vagas_prioritarias + [vaga for vaga in vagas if vaga not in vagas_prioritarias]

        if len(vagas_restantes) >= len(apartamentos):
            random.shuffle(vagas_restantes)
            for apartamento in apartamentos:
                vaga_selecionada = vagas_restantes.pop()
                Sorteio.objects.create(
                    apartamento=apartamento, 
                    vaga=vaga_selecionada
                )
        
        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('max_club')
    
    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'max_club/max_club.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })



# @staff_member_required
def max_club_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('max_club')
    else:
        return render(request, 'max_club/max_club_zerar.html')
    

def max_club_excel(request):
    caminho_modelo = 'static/assets/modelos/max_club.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'E{linha}'] = sorteio.vaga.vaga
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response


def max_club_qrcode(request):
    apartamentos_disponiveis = Apartamento.objects.all()  # Adiciona esta linha
    numero_apartamento = request.GET.get('apartamento')
    resultados_filtrados = None
    if numero_apartamento:
        resultados_filtrados = Sorteio.objects.filter(apartamento__numero_apartamento=numero_apartamento)
    
    return render(request, 'max_club/max_club_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,  # Certifique-se de adicionar esta linha
    })


# BIKE
@staff_member_required
def max_club_bike(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        SorteioBike.objects.all().delete()
        
        # Obter todos os apartamentos e grupos de vagas
        apartamentos = list(ApartamentoBike.objects.all())
        vagas = list(VagaBike.objects.all())

        # Embaralhar as vagas disponíveis
        random.shuffle(vagas)

        for apartamento in apartamentos:
            if vagas:
                vaga_selecionada = vagas.pop()
                SorteioBike.objects.create(
                    apartamento=apartamento, 
                    vaga=vaga_selecionada
                )
        
        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('max_club_bike')
    
    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = SorteioBike.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'max_club/max_club_bike.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })


@staff_member_required
def max_club_bike_zerar(request):
    if request.method == 'POST':
        SorteioBike.objects.all().delete()
        return redirect('max_club_bike')
    else:
        return render(request, 'max_club/max_club_bike_zerar.html')


def max_club_bike_excel(request):
    caminho_modelo = 'static/assets/modelos/max_club_bike.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = SorteioBike.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'E{linha}'] = sorteio.vaga.vaga
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio_bike.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response


def max_club_bike_qrcode(request):
    # Obter apartamentos relacionados à tabela SorteioBike
    apartamentos_disponiveis = ApartamentoBike.objects.filter(sorteiobike__isnull=False)

    numero_apartamento = request.GET.get('apartamento')
    resultados_filtrados = None

    if numero_apartamento:
        resultados_filtrados = SorteioBike.objects.filter(apartamento__numero_apartamento=numero_apartamento)
    
    return render(request, 'max_club/max_club_bike_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,  # Certifique-se de adicionar esta linha
    })

