from django.shortcuts import render, redirect
from .models import Apartamento, Vaga, Sorteio
from django.utils import timezone
import random
from django.contrib.admin.views.decorators import staff_member_required

# Excel
from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages


@staff_member_required
def porcelana_aleatorio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()
        
        # Obter todos os apartamentos e grupos de vagas
        apartamentos = list(Apartamento.objects.all())
        vagas = list(Vaga.objects.all())

        # Certifique-se de que existem vagas suficientes para todos os apartamentos
        if len(vagas) >= len(apartamentos):
            random.shuffle(vagas)

            for apartamento in apartamentos:
                vaga_selecionada = vagas.pop()
                Sorteio.objects.create(
                    apartamento=apartamento, 
                    vaga=vaga_selecionada
                )
        else:
           
            pass
        
        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('porcelana_aleatorio')
    
    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'porcelana/porcelana_aleatorio.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })


@staff_member_required
def zerar_porcelana(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('porcelana_aleatorio')
    else:
        return render(request, 'porcelana/porcelana_zerar.html')


def excel_porcelana(request):
    caminho_modelo = 'static/assets/modelos/sorteio_porcelana.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'E{linha}'] = sorteio.vaga.vaga
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response


def qrcode_porcelana(request):
    apartamentos_disponiveis = Apartamento.objects.all()  # Adiciona esta linha
    numero_apartamento = request.GET.get('apartamento')
    resultados_filtrados = None
    if numero_apartamento:
        resultados_filtrados = Sorteio.objects.filter(apartamento__numero_apartamento=numero_apartamento)
    
    return render(request, 'porcelana/porcelana_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,  # Certifique-se de adicionar esta linha
    })


def porcelana_inicio(request):
    	return render(request, 'porcelana/porcelana_inicio.html')

def porcelana_presenca(request):
    if request.method == 'POST':
        apartamento = Apartamento.objects.all()
        for item in apartamento:
            item.presenca = request.POST.get('presenca' + str(item.id)) == 'True'
            item.save()
        return redirect('porcelana_filtrar')  # Redireciona para a rota 'filtrar_presenca'
    apartamento = Apartamento.objects.all()
    return render(request, 'porcelana/porcelana_presenca.html', {"lista_de_presenca": apartamento})


def porcelana_filtrar(request):
    lista_de_presenca = Apartamento.objects.none()  # Inicia com uma queryset vazia
    if request.method == 'POST':
        lista_de_presenca = Apartamento.objects.all()  # Recupera todos os objetos quando o formulário é submetido
        if 'presentes' in request.POST:
            lista_de_presenca = lista_de_presenca.filter(presenca=True)
        if 'ausentes' in request.POST:
            lista_de_presenca = lista_de_presenca.filter(presenca=False)
    return render(request, 'porcelana/porcelana_filtrar.html', {"lista_de_presenca": lista_de_presenca})


def porcelana_s_apartamento(request):
    # Seleciona todos os apartamentos com presença confirmada e que ainda não têm vaga sorteada
    apartamentos_disponiveis = Apartamento.objects.filter(presenca=True).exclude(sorteio__apartamento__isnull=False)
    # Seleciona todas as vagas que ainda não estão em um sorteio
    vagas_disponiveis = Vaga.objects.exclude(sorteio__vaga__isnull=False)

    sorteio_finalizado = not apartamentos_disponiveis.exists() or not vagas_disponiveis.exists()
    item_de_presenca = None  # Inicializa a variável

    if request.method == 'POST':
        if 'realizar_sorteio' in request.POST and apartamentos_disponiveis.exists():
            # Sorteio de um apartamento aleatório
            apartamento_escolhido = random.choice(list(apartamentos_disponiveis))
            messages.success(request, f'Apartamento {apartamento_escolhido.numero_apartamento} foi selecionado para atribuição de vaga. Escolha a vaga agora.')

            # Armazena o apartamento escolhido para permitir seleção de vaga
            item_de_presenca = apartamento_escolhido

            # Renderiza a página com o apartamento sorteado para escolha da vaga
            return render(request, 'porcelana/porcelana_s_apartamento.html', {
                'sorteio_finalizado': sorteio_finalizado,
                'apartamentos_disponiveis': apartamentos_disponiveis,
                'vagas_disponiveis': vagas_disponiveis,
                'item_de_presenca': item_de_presenca
            })

        # Seção para escolher vaga para o apartamento selecionado
        if 'vaga_selecionada' in request.POST and 'apartamento_id' in request.POST:
            apartamento_id = request.POST.get('apartamento_id')
            vaga_selecionada = request.POST.get('vaga_selecionada')
            apartamento = Apartamento.objects.get(id=apartamento_id)
            vaga = Vaga.objects.get(vaga=vaga_selecionada)

            # Cria um novo registro de sorteio associando o apartamento à vaga escolhida
            novo_sorteio = Sorteio(apartamento=apartamento, vaga=vaga)
            novo_sorteio.save()

            messages.success(request, f'Vaga {vaga.vaga} confirmada para o apartamento {apartamento.numero_apartamento} com sucesso!')
            return redirect('porcelana_s_apartamento')

    return render(request, 'porcelana/porcelana_s_apartamento.html', {
        'sorteio_finalizado': sorteio_finalizado,
        'apartamentos_disponiveis': apartamentos_disponiveis,
        'vagas_disponiveis': vagas_disponiveis,
        'item_de_presenca': item_de_presenca
    })


def porcelana_sorteio(request):
    	return render(request, 'porcelana/porcelana_sorteio.html')


# @staff_member_required
# def porcelana_final(request):
#     if request.method == 'POST':
#         # Limpar registros anteriores de sorteio
#         Sorteio.objects.all().delete()
        
#         # Obter todos os apartamentos e grupos de vagas
#         apartamentos = list(Apartamento.objects.all())
#         vagas = list(Vaga.objects.all())

#         # Certifique-se de que existem vagas suficientes para todos os apartamentos
#         if len(vagas) >= len(apartamentos):
#             random.shuffle(vagas)

#             for apartamento in apartamentos:
#                 vaga_selecionada = vagas.pop()
#                 Sorteio.objects.create(
#                     apartamento=apartamento, 
#                     vaga=vaga_selecionada
#                 )
#         else:
           
#             pass
        
#         # Armazenar informações do sorteio na sessão
#         request.session['sorteio_iniciado_nc'] = True
#         request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

#         return redirect('porcelana_aleatorio')
    
#     else:
#         sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
#         resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
#         vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

#         return render(request, 'porcelana/porcelana_aleatorio.html', {
#             'resultados_sorteio_nc': resultados_sorteio_nc,
#             'vagas_atribuidas_nc': vagas_atribuidas_nc,
#             'sorteio_iniciado_nc': sorteio_iniciado_nc,
#             'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
#         })

from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from .models import Apartamento, Vaga, Sorteio
import random

@staff_member_required
def porcelana_final(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio pode não ser necessário se você quer manter histórico
        # Sorteio.objects.all().delete()

        # Obter todos os apartamentos que não estão atualmente associados a um sorteio
        apartamentos = list(Apartamento.objects.exclude(sorteio__isnull=True))
        vagas = list(Vaga.objects.exclude(sorteio__isnull=True))

        # Certifique-se de que existem vagas suficientes para todos os apartamentos
        if len(vagas) >= len(apartamentos):
            random.shuffle(vagas)

            for apartamento in apartamentos:
                if not apartamento.sorteio_set.exists():  # Verifica se o apartamento já foi sorteado
                    vaga_selecionada = vagas.pop()
                    Sorteio.objects.create(
                        apartamento=apartamento,
                        vaga=vaga_selecionada
                    )
        else:
            # Tratar a situação onde não há vagas suficientes para os apartamentos
            messages.error(request, "Não há vagas suficientes para todos os apartamentos.")

        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('porcelana_aleatorio')

    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'porcelana/porcelana_aleatorio.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })
