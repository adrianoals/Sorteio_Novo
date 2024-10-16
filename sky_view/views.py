from django.shortcuts import render, redirect
from django.utils import timezone
import random
from django.http import HttpResponse
from django.urls import reverse
# from openpyxl import Workbook  # Para gerar o Excel
from openpyxl import load_workbook
from io import BytesIO  # Para manipular imagens em memória
from .models import Apartamento, Vaga, Sorteio


# View para realizar o sorteio
def sky_view_sorteio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio (opcional)
        Sorteio.objects.all().delete()

        # Obtenha todos os apartamentos e vagas disponíveis
        apartamentos = Apartamento.objects.all()
        vagas_simples = list(Vaga.objects.filter(tipo_vaga='simples'))  # Convertido para lista
        vagas_duplas = list(Vaga.objects.filter(tipo_vaga='dupla'))  # Convertido para lista

        resultados_sorteio = []

        # Loop por cada apartamento e sortear vagas
        for apartamento in apartamentos:
            if apartamento.direito_vaga_dupla:
                # Se o apartamento tiver direito a vaga dupla, sorteia uma vaga dupla
                if vagas_duplas:
                    vaga_dupla = random.choice(vagas_duplas)
                    sorteio = Sorteio.objects.create(apartamento=apartamento, vaga=vaga_dupla)
                    resultados_sorteio.append(sorteio)
                    # Remover a vaga dupla sorteada da lista
                    vagas_duplas.remove(vaga_dupla)
            elif apartamento.direito_duas_vagas_livres:
                # Para apartamentos com duas vagas livres (sequenciais no mesmo subsolo)
                subsolo = random.choice(['1º Subsolo', '2º Subsolo', '3º Subsolo', '4º Subsolo'])
                vagas_no_subsolo = [v for v in vagas_simples if v.subsolo == subsolo]
                if len(vagas_no_subsolo) >= 2:
                    vaga1 = vagas_no_subsolo[0]
                    vaga2 = vagas_no_subsolo[1]
                    Sorteio.objects.create(apartamento=apartamento, vaga=vaga1)
                    Sorteio.objects.create(apartamento=apartamento, vaga=vaga2)
                    # Remover as vagas simples sorteadas da lista
                    vagas_simples.remove(vaga1)
                    vagas_simples.remove(vaga2)
            else:
                # Sorteia uma vaga simples para outros apartamentos
                if vagas_simples:
                    vaga_simples = random.choice(vagas_simples)
                    sorteio = Sorteio.objects.create(apartamento=apartamento, vaga=vaga_simples)
                    resultados_sorteio.append(sorteio)
                    # Remover a vaga simples sorteada da lista
                    vagas_simples.remove(vaga_simples)

        # Marcar o sorteio como iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        # Redireciona para a mesma página após o sorteio
        return redirect(reverse('sky_view_sorteio'))

    # Se o método for GET, exibe os resultados ou a interface de sorteio
    sorteio_iniciado = request.session.get('sorteio_iniciado', False)
    vagas_atribuidas = Sorteio.objects.exists()
    resultados_sorteio = Sorteio.objects.all() if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('horario_conclusao', '')  # Exibe o horário de conclusão
    }

    return render(request, 'sky_view/sky_view_sorteio.html', context)


def sky_view_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteio_skyview.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['B8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'B{linha}'] = sorteio.apartamento.numero  # Número do apartamento
        ws[f'C{linha}'] = sorteio.vaga.numero  # Número da vaga
        ws[f'D{linha}'] = f'Subsolo {sorteio.vaga.subsolo}'  # Subsolo
        ws[f'E{linha}'] = sorteio.vaga.tipo_vaga  # Tipo da vaga
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sorteio_sky_view.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response



def sky_view_qrcode(request):
    # Obter todos os apartamentos para preencher o dropdown
    apartamentos_disponiveis = Apartamento.objects.all()
    
    # Obter o apartamento selecionado através do filtro (via GET)
    numero_apartamento = request.GET.get('apartamento')

    # Inicializar a variável de resultados filtrados como uma lista vazia
    resultados_filtrados = []

    # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
    if numero_apartamento:
        # Buscar os sorteios para o apartamento selecionado
        resultados_filtrados = Sorteio.objects.filter(apartamento__numero=numero_apartamento)

    return render(request, 'sky_view/sky_view_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })



def sky_view_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('sky_view_sorteio')
    else:
        return render(request, 'sky_view/sky_view_zerar.html')
    

