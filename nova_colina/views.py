from django.shortcuts import render

# def nova_colina(request):
#     	return render(request, 'nova_colina/nova_colina.html')


from django.shortcuts import render, redirect
# from .models import Apartamento, VagaSimples, VagaDupla, Sorteio
from django.utils import timezone
import random

# def nova_colina(request):
#     if request.method == 'POST':
#         # Limpar registros anteriores
#         Sorteio.objects.all().delete()
        
#         # Obter todos os apartamentos
#         apartamentos = list(Apartamento.objects.all())

#         # Obter todas as vagas simples e duplas disponíveis
#         vagas_simples = list(VagaSimples.objects.all())
#         vagas_duplas = list(VagaDupla.objects.all())

#         # Certifique-se de que existem vagas suficientes para todos os apartamentos
#         if len(vagas_simples) >= len(apartamentos) and len(vagas_duplas) >= len(apartamentos):
#             # Embaralhar as listas de vagas
#             random.shuffle(vagas_simples)
#             random.shuffle(vagas_duplas)

#             # Iterar sobre os apartamentos e atribuir uma vaga simples e uma vaga dupla a cada um
#             for apartamento in apartamentos:
#                 vaga_simples_selecionada = vagas_simples.pop()  # Remover uma vaga simples da lista
#                 vaga_dupla_selecionada = vagas_duplas.pop()  # Remover uma vaga dupla da lista
#                 Sorteio.objects.create(
#                     apartamento=apartamento, 
#                     vaga_simples=vaga_simples_selecionada, 
#                     vaga_dupla=vaga_dupla_selecionada
#                 )
#         else:
#             # Tratar o caso em que não há vagas suficientes para todos os apartamentos
#             # Você pode definir uma mensagem de erro e retorná-la ao template, se desejado
#             pass
        
#         # Marcar na sessão que o sorteio foi iniciado e armazenar o horário de conclusão
#         request.session['sorteio_iniciado_nc'] = True
#         request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

#         # Redirecionar para a mesma página para mostrar os resultados do sorteio
#         return redirect('nova_colina')
    
#     else:
#         # Verificar se o sorteio já foi iniciado
#         sorteio_iniciado_nc = request.session.get('sorteio_iniciado', False)
#         resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga_simples', 'vaga_dupla').all()
#         vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

#         return render(request, 'nova_colina/nova_colina.html', {
#             'resultados_sorteio_nc': resultados_sorteio_nc,
#             'vagas_atribuidas_nc': vagas_atribuidas_nc,
#             'sorteio_iniciado_nc': sorteio_iniciado_nc,
#             'horario_conclusao_nc': request.session.get('horario_conclusao', '')
#         })


# from django.shortcuts import render, redirect
# from .models import Apartamento, GrupoVaga, Sorteio
# from django.utils import timezone
# import random

# def nova_colina(request):
#     if request.method == 'POST':
#         # Limpar registros anteriores de sorteio
#         Sorteio.objects.all().delete()
        
#         # Obter todos os apartamentos e grupos de vagas
#         apartamentos = list(Apartamento.objects.all())
#         grupos_vagas = list(GrupoVaga.objects.all())

#         # Certifique-se de que existem grupos de vagas suficientes para todos os apartamentos
#         if len(grupos_vagas) >= len(apartamentos):
#             # Embaralhar as listas de grupos de vagas
#             random.shuffle(grupos_vagas)

#             # Iterar sobre os apartamentos e atribuir um grupo de vagas a cada um
#             for apartamento in apartamentos:
#                 grupo_vaga_selecionado = grupos_vagas.pop()  # Remover um grupo de vagas da lista
#                 Sorteio.objects.create(
#                     apartamento=apartamento, 
#                     vagas=grupo_vaga_selecionado
#                 )
#         else:
#             # Tratar o caso em que não há grupos de vagas suficientes para todos os apartamentos
#             # Você pode definir uma mensagem de erro e retorná-la ao template, se desejado
#             pass
        
#         # Marcar na sessão que o sorteio foi iniciado e armazenar o horário de conclusão
#         request.session['sorteio_iniciado_nc'] = True
#         request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

#         # Redirecionar para a mesma página para mostrar os resultados do sorteio
#         return redirect('nova_colina')
    
#     else:
#         # Verificar se o sorteio já foi iniciado
#         sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
#         resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vagas').all()
#         vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

#         return render(request, 'nova_colina/nova_colina.html', {
#             'resultados_sorteio_nc': resultados_sorteio_nc,
#             'vagas_atribuidas_nc': vagas_atribuidas_nc,
#             'sorteio_iniciado_nc': sorteio_iniciado_nc,
#             'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
#         })


from django.shortcuts import render, redirect
from .models import Apartamento, GrupoVaga, Sorteio
from django.utils import timezone
import random

def nova_colina(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()
        
        # Predefinições de apartamentos e suas vagas correspondentes
        predefinicoes = {
            'Apto 901': 'Vaga 01, Vaga 10 e Vaga 11',
            'Apto 903': 'Vaga 02, Vaga 12 e Vaga 13',
            'Apto 1201': 'Vaga 44, Vaga 45 e Vaga 46',
            'Apto 904': 'Vaga 49, Vaga 50 e Vaga 51',
            'Apto 1302': 'Vaga 81, Vaga 72 e Vaga 73',
            'Apto 1304': 'Vaga 82, Vaga 74 e Vaga 75',
            'Apto 303': 'Vaga 76, Vaga 77 e Vaga 78',
            'Apto 104': 'Vaga 122, Vaga 123 e Vaga 124',
            'Apto 403': 'Vaga 16, Vaga 162 e Vaga 163',
        }

        # Atribuir as vagas predefinidas
        for apto, vagas in predefinicoes.items():
            apartamento = Apartamento.objects.get(numero_apartamento=apto)
            grupo_vaga = GrupoVaga.objects.get(vagas=vagas)
            Sorteio.objects.create(apartamento=apartamento, vagas=grupo_vaga)

        # Obter todos os apartamentos e grupos de vagas, excluindo os predefinidos
        apartamentos = list(Apartamento.objects.exclude(numero_apartamento__in=predefinicoes.keys()))
        grupos_vagas = list(GrupoVaga.objects.exclude(vagas__in=predefinicoes.values()))

        # Certifique-se de que existem grupos de vagas suficientes para os apartamentos restantes
        if len(grupos_vagas) >= len(apartamentos):
            random.shuffle(grupos_vagas)

            for apartamento in apartamentos:
                grupo_vaga_selecionado = grupos_vagas.pop()
                Sorteio.objects.create(
                    apartamento=apartamento, 
                    vagas=grupo_vaga_selecionado
                )
        else:
            # Tratar o caso em que não há grupos de vagas suficientes para os apartamentos restantes
            pass
        
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('nova_colina')
    
    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        # Ordenar os resultados do sorteio pelo ID do apartamento
        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vagas').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'nova_colina/nova_colina.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })


def zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('nova_colina')
    else:
        return render(request, 'nova_colina/zerar.html')

from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone
from .models import Sorteio

def excel_nova_colina(request):
    caminho_modelo = 'static/assets/modelos/sorteio_novo2.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vagas').order_by('apartamento__id').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'E{linha}'] = sorteio.vagas.vagas
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response


# from openpyxl import load_workbook
# from django.http import HttpResponse
# from django.utils import timezone
# from .models import Sorteio

# def excel_nova_colina(request):
#     # Construir o caminho completo para o arquivo modelo
#     caminho_modelo = 'static/assets/modelos/sorteio_novo2.xlsx' 

#     # Carregar o workbook do modelo
#     wb = load_workbook(caminho_modelo)
#     ws = wb.active

#     # Obter os resultados do sorteio
#     resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vagas').all()

#     # Recuperar o horário de conclusão do sorteio da sessão
#     horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')

#     # Escrever o horário de conclusão na célula C8
#     ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

#     # Supondo que você queira começar a inserir os dados a partir da linha 10
#     linha = 10
#     for sorteio in resultados_sorteio_nc:
#         ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
#         ws[f'E{linha}'] = sorteio.vagas.vagas
#         linha += 1

#     # Preparar a resposta para enviar o arquivo
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     nome_arquivo = "resultado_sorteio.xlsx"
#     response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

#     # Salvar o workbook modificado no response
#     wb.save(response)

#     return response


def filtro_apartamento(request):
    apartamentos_disponiveis = Apartamento.objects.all()  # Adiciona esta linha
    numero_apartamento = request.GET.get('apartamento')
    resultados_filtrados = None
    if numero_apartamento:
        resultados_filtrados = Sorteio.objects.filter(apartamento__numero_apartamento=numero_apartamento)
    
    return render(request, 'nova_colina/nv_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,  # Certifique-se de adicionar esta linha
    })
